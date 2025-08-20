# -*- coding: utf-8 -*-
# parse_cedolini.py — Estrazione cedolini PDF -> Dettaglio + Aggregati + Anagrafica (XLSX) con dropdown 'sede_operativa'
# Robustezza: ancore LUL/Datev, split chunk “validi”, regex a fine riga (no catture a fiume),
# EU->float, lineage source_file, anno con fallback (pagina o nome file), QA check basilari.
import re, argparse, sys, time
from pathlib import Path
import pandas as pd
from pdfminer.high_level import extract_text
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- costanti & util ----------
CONTROL_CHARS = re.compile(r"[\000-\010\013\014\016-\037]")
CF_RE = r"\b([A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z])\b"
EU_NUM = r"[-+]?\d{1,3}(?:\.\d{3})*(?:,\d+)?|[-+]?\d+(?:,\d+)?"

PRES_CODES = ["FE","FT","A1","RS","ROL","MAL","INF","MAT","PERM","STRAORD","FG"]  # FG=festività non goduta (presenti nei tuoi PDF)

MONTHS_MAP = {
    "gen": "Gennaio", "feb": "Febbraio", "mar": "Marzo", "apr": "Aprile",
    "mag": "Maggio", "giu": "Giugno", "lug": "Luglio", "ago": "Agosto",
    "set": "Settembre", "ott": "Ottobre", "nov": "Novembre", "dic": "Dicembre"
}

def euro_to_float(s):
    if not s: return None
    s = str(s).replace('\xa0','').replace(' ', '')
    s = s.replace('.', '').replace(',', '.')
    m = re.findall(r"[-+]?\d+(?:\.\d+)?", s)
    return float(m[-1]) if m else None

def clean_text(txt: str) -> str:
    # pdfminer può produrre spazi “artistici”: normalizzo ma conservo le maiuscole utili
    txt = txt.replace("\r", "")
    return CONTROL_CHARS.sub("", txt)

def find_first_line(text, patterns, group=1, flags=re.IGNORECASE):
    # cerca su base “fine riga” (no DOTALL) per evitare colate di testo
    for pat in patterns:
        m = re.search(pat, text, flags)
        if m:
            return m.group(group).strip()
    return None

def parse_month(text):
    m = find_first_line(text, [r"\bMESE\s+RETRIBUITO\s+([A-Za-zÀ-ÖØ-öø-ÿ]+)\b"])
    if m: return m
    # fallback: individua pattern “Aprile 2025” nel frontespizio
    m2 = find_first_line(text, [r"\b(Gennaio|Febbraio|Marzo|Aprile|Maggio|Giugno|Luglio|Agosto|Settembre|Ottobre|Novembre|Dicembre)\b"])
    return m2

def parse_year(text, source_name):
    y = find_first_line(text, [r"\bMESE\s+RETRIBUITO\s+[A-Za-zÀ-ÖØ-öø-ÿ]+\s+(20\d{2})\b",
                               r"\bANNO\s+(20\d{2})\b",
                               r"\b(20\d{2})\b"])  # ultimo fallback nella pagina
    if y: return y
    # ultimissimo fallback: numero a 4 cifre nel nome file
    m = re.search(r"(20\d{2})", source_name)
    return m.group(1) if m else None

def month_matches(text, month_key):
    if not month_key: return [False]*len(text)
    k = month_key.strip().lower()[:3]
    return text.str.contains(k, case=False, na=False)

# ---------- splitting & parsing ----------
def split_employees(text):
    # spezza su “DIPENDENTE ” e filtra chunk “validi” (devono contenere CF o la parola QUALIFICA)
    raw_parts = re.split(r"(?=DIPENDENTE\s)", text)
    parts = []
    for p in raw_parts:
        if re.search(CF_RE, p) or re.search(r"\bQUALIFICA\b", p, re.IGNORECASE):
            parts.append(p)
    return parts

def parse_chunk(chunk, comp, verbose=False):
    get = lambda pats: find_first_line(chunk, pats)

    rec = {
        # metadati azienda + periodo
        "azienda_denominazione": comp.get("azienda_denominazione"),
        "azienda_cf": comp.get("azienda_cf"),
        "azienda_piva": comp.get("azienda_piva"),
        "mese_retribuito": comp.get("mese_retribuito"),
        "anno": comp.get("anno"),
        "source_file": comp.get("source_file"),
        # colonna manuale a dropdown
        "sede_operativa": None,
    }

    # --- anagrafica dipendente / contrattuale (fine riga, no DOTALL)
    name = get([r"DIPENDENTE\s+([^\n]+?)\s+(?:QUALIFICA|CODICE\s+FISCALE|RETRIBUZIONE|TIPO|MANSIONE)\b"])
    if not name:
        name = get([r"DIPENDENTE\s+([^\n]+)"])
    rec["dipendente_nome"] = " ".join((name or "").split()) or None

    rec["dipendente_cf"] = get([CF_RE])
    rec["matricola_inps"] = get([r"\bMATRICOLA\s+INPS\s+([0-9/]+)\b"])
    rec["qualifica"] = get([r"\bQUALIFICA\s+([^\n]+)"])
    rec["mansione"] = get([r"\bMANSIONE\s+([^\n]+)"])
    rec["livello"] = get([r"\bLIVELLO\s+([^\n]+)"])
    rec["tipo_rapporto"] = get([r"\bTIPO\s+RAPPORTO\s+([^\n]+)"])
    rec["data_assunzione"] = get([r"\bDATA\s+ASSUNZIONE\s+(\d{1,2}/\d{1,2}/\d{4})\b"])
    rec["data_cessazione"] = get([r"\bDATA\s+CESSAZIONE\s+(\d{1,2}/\d{1,2}/\d{4})\b"])

    # --- retribuzione variabili / fisse (token tipici LUL Datev)
    rec["totale_competenze"] = euro_to_float(get([rf"\bTOTALE\s+COMPETENZE[^\d]*({EU_NUM})"]))
    rec["totale_trattenute"] = euro_to_float(get([rf"\bTOTALE\s+RITENUTE[^\d]*({EU_NUM})", rf"\bTOTALE\s+TRATTENUTE[^\d]*({EU_NUM})"]))
    rec["netto_a_pagare"] = euro_to_float(get([rf"\bNETTO\s*(?:IN\s*BUSTA|A\s*PAGARE)[^\d]*({EU_NUM})"]))
    rec["imponibile_previdenziale"] = euro_to_float(get([rf"\bTOTALE\s+IMPONIBILE\s+INPS[^\d]*({EU_NUM})", rf"\bIMPO[NM]IBILE\s+PREVIDENZIALE[^\d]*({EU_NUM})"]))
    rec["imponibile_fiscale"] = euro_to_float(get([rf"\bIMPO[NM]IBILE\s+FISCALE[^\d]*({EU_NUM})"]))
    rec["inps_dip"] = euro_to_float(get([rf"\bRITENUTE\s+INPS[^\d]*({EU_NUM})", rf"\bCONTRIB(?:\.)?\s*INPS\s+DIP\.*[^\d]*({EU_NUM})"]))
    rec["inps_azienda"] = euro_to_float(get([rf"(?:\bINPS\s+DITTA|\bCONTRIB(?:\.)?\s*INPS\s+DITTA)[^\d]*({EU_NUM})"]))
    rec["inail_azienda"] = euro_to_float(get([rf"\bINAIL[^\d]*({EU_NUM})"]))
    rec["tfr_mese"] = euro_to_float(get([rf"\bTFR\s+DEL\s+MESE[^\d]*({EU_NUM})"]))
    rec["quota_anno_tfr"] = euro_to_float(get([rf"\bRIVALUTAZIONE\s+QUOTA\s+ANNO\s+TFR[^\d]*({EU_NUM})",
                                               rf"\bQUOTA\s+ANNO\s+TFR[^\d]*({EU_NUM})"]))
    # spesso “costo azienda” non è stampato: lo calcoliamo se i pezzi ci sono
    costo_az = euro_to_float(get([rf"\bCOSTO\s+AZIENDA[^\d]*({EU_NUM})"]))
    if costo_az is None:
        parts = [rec["totale_competenze"], rec["inps_azienda"], rec["inail_azienda"], rec["tfr_mese"]]
        if any(x is not None for x in parts):
            costo_az = sum(x or 0.0 for x in parts)
    rec["costo_azienda"] = costo_az

    # --- presenze: conta codici (upper per sicurezza)
    up = chunk.upper()
    for code in PRES_CODES:
        rec[f"occ_{code}"] = len(re.findall(rf"\b{code}\b", up))

    # scarto chunk “fantasma”: servono almeno CF o nome + qualifica
    if not rec["dipendente_cf"] and not (rec["dipendente_nome"] and (rec["qualifica"] or rec["mansione"])):
        return None

    if verbose:
        who = rec.get("dipendente_nome") or rec.get("dipendente_cf")
        print(f"   → {who} | competenze={rec['totale_competenze']} netto={rec['netto_a_pagare']}")
    return rec

def parse_pdf_to_records(path: Path, verbose=False):
    txt = extract_text(str(path)) or ""
    txt = clean_text(txt)
    comp = {
        "azienda_denominazione": find_first_line(txt, [r"\bAZIENDA\s+([^\n]+)"]),
        "azienda_cf": find_first_line(txt, [r"\bCODICE\s+FISCALE\s+([0-9/]{5,})"]),
        "azienda_piva": find_first_line(txt, [r"\bPARTITA\s+IVA\s+([0-9/]{5,})"]),
        "mese_retribuito": parse_month(txt),
        "anno": parse_year(txt, path.name),
        "source_file": path.name,
    }
    chunks = split_employees(txt)
    if verbose: print(f"[{path.name}] blocchi validi: {len(chunks)} | mese={comp['mese_retribuito']} anno={comp['anno']}")
    recs = []
    for ch in chunks:
        r = parse_chunk(ch, comp, verbose=verbose)
        if r: recs.append(r)
    return recs

# ---------- sanitizzazione & QA ----------
def sanitize_df(df):
    df2 = df.copy()
    for c in df2.columns:
        if df2[c].dtype == object:
            df2[c] = df2[c].astype(str).map(lambda x: CONTROL_CHARS.sub("", x))
    # normalizza capitalizzazione nome
    if "dipendente_nome" in df2.columns:
        df2["dipendente_nome"] = df2["dipendente_nome"].replace({"": None, "None": None})
        df2["dipendente_nome"] = df2["dipendente_nome"].dropna().map(str.title).reindex(df2.index, fill_value=None)
    return df2

def qa_checks(df):
    out = []
    # 1) competenze - trattenute ≈ netto (±1 euro per arrotondamenti)
    if {"totale_competenze","totale_trattenute","netto_a_pagare"}.issubset(df.columns):
        delta = (df["totale_competenze"].fillna(0)
                 - df["totale_trattenute"].fillna(0)
                 - df["netto_a_pagare"].fillna(0)).abs()
        n_bad = int((delta > 1.0).sum())
        if n_bad:
            out.append(f"[QA] Mismatch competenze-trattenute-netto: {n_bad} righe > 1€")
    # 2) imponibile INPS non deve superare le competenze utili (check soft)
    if {"imponibile_previdenziale","totale_competenze"}.issubset(df.columns):
        n_bad = int((df["imponibile_previdenziale"] > df["totale_competenze"]).sum())
        if n_bad:
            out.append(f"[QA] Imponibile INPS > competenze in {n_bad} righe")
    # 3) duplicati CF×mese×anno×file
    if {"dipendente_cf","mese_retribuito","anno","source_file"}.issubset(df.columns):
        dup = df.duplicated(subset=["dipendente_cf","mese_retribuito","anno","source_file"], keep=False)
        n_dup = int(dup.sum())
        if n_dup:
            out.append(f"[QA] Duplicati CF×mese×anno×file: {n_dup}")
    # 4) mese/anno mancanti
    if "anno" in df.columns:
        miss_y = int(df["anno"].isna().sum())
        if miss_y: out.append(f"[QA] Anno mancante su {miss_y} righe")
    if "mese_retribuito" in df.columns:
        miss_m = int(df["mese_retribuito"].isna().sum())
        if miss_m: out.append(f"[QA] Mese mancante su {miss_m} righe")
    return out

def write_excel_with_dropdown(df_det, df_agg, df_anag, out_xlsx):
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xlw:
        df_det.to_excel(xlw, sheet_name="DettaglioCedolini", index=False)
        df_agg.to_excel(xlw, sheet_name="AggregatiMensili", index=False)
        df_anag.to_excel(xlw, sheet_name="Anagrafica", index=False)
        wb = xlw.book
        ws = wb["DettaglioCedolini"]
        # dropdown sede_operativa
        try:
            col_idx = list(df_det.columns).index("sede_operativa") + 1
            dv = DataValidation(
                type="list",
                formula1='"Garibaldi,Pompea,Entrambi"',
                allow_blank=True,
                showErrorMessage=True
            )
            ws.add_data_validation(dv)
            last_row = ws.max_row
            col_letter = get_column_letter(col_idx)
            dv.add(f"{col_letter}2:{col_letter}{last_row}")
        except ValueError:
            # se la colonna non esiste, non fa nulla
            pass

def build_outputs(pdf_paths, out_xlsx, verbose=False):
    # Parse tutti i PDF
    records = []
    for i, p in enumerate(pdf_paths, 1):
        if verbose: print(f"[{i}/{len(pdf_paths)}] {p.name}")
        records += parse_pdf_to_records(p, verbose=verbose)
    df = pd.DataFrame(records)
    if df.empty:
        raise SystemExit("Nessun record estratto: controlla i PDF o le regex.")

    # Tipizzazioni numeriche
    num_cols = ["totale_competenze","totale_trattenute","netto_a_pagare","imponibile_fiscale",
                "imponibile_previdenziale","inps_dip","inps_azienda","inail_azienda",
                "tfr_mese","quota_anno_tfr","costo_azienda"] + [f"occ_{c}" for c in PRES_CODES]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # Chiave tecnica
    df["key_dipendente"] = df["dipendente_cf"].where(
        df["dipendente_cf"].notna(),
        (df["dipendente_nome"].fillna("")+"|"+df["mese_retribuito"].fillna("")+"|"+df["anno"].fillna("").astype(str))
    )

    # Anagrafica storica
    df_anag = (df.sort_values(["dipendente_cf","data_assunzione"])
                 .groupby("key_dipendente", dropna=False)
                 .agg(dipendente_nome=("dipendente_nome","last"),
                      dipendente_cf=("dipendente_cf","last"),
                      matricola_inps=("matricola_inps","last"),
                      qualifica=("qualifica","last"),
                      mansione=("mansione","last"),
                      livello=("livello","last"),
                      tipo_rapporto=("tipo_rapporto","last"),
                      data_assunzione=("data_assunzione","min"),
                      data_cessazione=("data_cessazione","max"))
                 .reset_index(drop=True))

    # Aggregati mensili per CFO (per sede_operativa manuale)
    group_keys = ["key_dipendente","dipendente_nome","dipendente_cf","mese_retribuito","anno","sede_operativa"]
    df_mensile = (df.groupby(group_keys, dropna=False)[num_cols]
                    .sum(min_count=1)
                    .reset_index())

    # Sanitize + QA
    df_det_s = sanitize_df(df)
    df_agg_s = sanitize_df(df_mensile)
    df_anag_s = sanitize_df(df_anag)
    msgs = qa_checks(df_det_s)
    if msgs:
        print("\n".join(msgs))

    # Scrittura Excel con dropdown
    write_excel_with_dropdown(df_det_s, df_agg_s, df_anag_s, out_xlsx)
    print(f"[OK] Creato: {out_xlsx}")
    return out_xlsx

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cartella", required=True, help="Cartella con cedolini PDF")
    ap.add_argument("--out", required=True, help="Percorso XLSX di output")
    ap.add_argument("--verbose", action="store_true", help="Log di avanzamento parsing")
    args = ap.parse_args()

    pdfs = list(Path(args.cartella).glob("*.pdf"))
    if not pdfs:
        raise SystemExit("Nessun PDF trovato nella cartella")

    build_outputs(pdfs, args.out, verbose=args.verbose)

if __name__ == "__main__":
    main()

